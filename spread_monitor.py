# -*- coding: utf-8 -*-
"""
期货相邻月份价差监控 (基于 TqSdk)

方向统一 近月-远月。三腿：近月 outright / 远月 outright / 交易所跨期价差合约。
合成价差(由两腿盘口)与交易所组合并成一个合并盘口，叠加套利筛选视图：
  标签：主力/注销月/注销月后第一个月/近月（多标签全显示）
  合并买价 = max(合成卖价差, 组合买价)   （买一：能卖出价差的价）
  合并卖价 = min(合成买价差, 组合卖价)   （卖一：能买入价差的价）
  最新价差 = 合成最新(近last-远last) 夹在 [合并买卖价] 之间
  价差比例%   = 最新价差 / 近月最新 × 100
  仓储费比例   = 最新价差 / (−仓储费 × 30)
  转抛比例     = 最新价差 / (−1.1 × 年化资金利率% / 跨期月份差 − 仓储费 × 30)
HTML 支持点列头排序、多条件筛选、资金利率输入框(实时重算转抛比例)，状态本地保存、可暂停自动刷新；Excel 开自动筛选。
"""

import os
import ast
import math
import time
import json
import html
import configparser
from datetime import datetime

import pandas as pd
from tqsdk import TqApi, TqAuth

FUTURE_EXCHANGES = ["SHFE", "DCE", "CZCE", "INE", "GFEX", "CFFEX"]
COMBINE_EXCHANGES = ["DCE", "CZCE"]
TODAY_YEAR = datetime.now().year


# --------------------------------------------------------------------------- #
# 工具
# --------------------------------------------------------------------------- #
def is_nan(x):
    return x is None or (isinstance(x, float) and math.isnan(x))


def clean(x):
    return None if is_nan(x) else x


def code_of(symbol):
    return symbol.split(".", 1)[1] if "." in symbol else symbol


_F_RE = None


def is_settle_f(symbol):
    """F结尾的结算价/新标准品合约(如 l2607F)：全系统一律排除（用户决定）。"""
    global _F_RE
    if _F_RE is None:
        import re
        _F_RE = re.compile(r"\d[Ff]$")
    return bool(_F_RE.search(code_of(symbol)))


def product_of(symbol):
    return "".join(ch for ch in code_of(symbol) if not ch.isdigit())


def month_key(symbol):
    digits = "".join(ch for ch in code_of(symbol) if ch.isdigit())
    if len(digits) == 4:
        year = 2000 + int(digits[:2]); mm = int(digits[2:])
    elif len(digits) == 3:
        d = int(digits[0]); mm = int(digits[1:])
        decade = (TODAY_YEAR // 10) * 10
        year = decade + d
        if year < TODAY_YEAR - 1:
            year += 10
    else:
        return 999999
    return year * 100 + mm


def month_diff(near, far):
    """跨期月份差(远-近)，按自然月计。"""
    a, b = month_key(near), month_key(far)
    if a >= 999999 or b >= 999999:
        return None
    return (b // 100 * 12 + b % 100) - (a // 100 * 12 + a % 100)


def parse_combine(symbol):
    try:
        ex, rest = symbol.split(".", 1)
        prefix, legs = rest.split(" ", 1)
        l1, l2 = legs.split("&")
        f1, f2 = f"{ex}.{l1.strip()}", f"{ex}.{l2.strip()}"
        if is_settle_f(f1) or is_settle_f(f2):
            return None            # F结算价合约组合一律排除
        if product_of(f1) != product_of(f2):
            return None
        return ex, prefix, f1, f2
    except Exception:
        return None


# --------------------------------------------------------------------------- #
# 品种配置
# --------------------------------------------------------------------------- #
def load_product_meta(path):
    meta = {}
    if not path or not os.path.exists(path):
        print(f"[警告] 未找到品种配置 {path}，转抛/仓储费/标签将留空")
        return meta
    import openpyxl
    ws = openpyxl.load_workbook(path, data_only=True)["Sheet1"]
    h = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    for r in range(2, ws.max_row + 1):
        p = ws.cell(r, h["product"]).value
        if not p:
            continue
        fee = ws.cell(r, h["daily_storage_fee"]).value
        dm = ws.cell(r, h["delivery_month"]).value
        roll = ws.cell(r, h["roll_over_arbitrage"]).value
        try:
            cancel = set(int(x) for x in ast.literal_eval(str(dm).replace("，", ",")))
        except Exception:
            cancel = set()
        decay = []
        if "time_decay" in h:
            tv = ws.cell(r, h["time_decay"]).value
            if tv and str(tv).strip():
                try:
                    decay = json.loads(str(tv).replace("，", ","))
                except Exception:
                    print(f"[警告] {p} time_decay 解析失败: {tv}")
        ind = ws.cell(r, h["industry_name"]).value if "industry_name" in h else None
        meta[str(p).upper()] = {
            "roll": (str(roll).strip() if roll is not None else ""),
            "fee": (float(fee) if isinstance(fee, (int, float)) else None),
            "cancel": cancel,
            "decay": decay,
            "industry": (str(ind).strip() if ind else ""),
        }
    return meta


INDEX_PRODUCTS = {"IF", "IC", "IH", "IM"}   # 股指：交割月正常交易


def _next_month(ym):
    y, m = ym // 100, ym % 100
    return (y + 1) * 100 + 1 if m == 12 else ym + 1


def can_roll(near_month, far_month, cancel_set):
    """月差可转抛 = 区间 [近月, 远月) 内无注销月（近月含，远月为退出点不含）。"""
    if not cancel_set:
        return True
    m = near_month
    while m < far_month:
        if (m % 100) in cancel_set:
            return False
        m = _next_month(m)
    return True


def in_delivery_month(near_month, product):
    """近月是否已进交割月（当前月>=合约月）；股指例外。"""
    if str(product).upper() in INDEX_PRODUCTS:
        return False
    cur = datetime.now().year * 100 + datetime.now().month
    return near_month is not None and near_month <= cur


def decay_cost(near_month, far_month, windows, delivery_day=15):
    """特殊时间贴水成本(元/吨)：持有期[近月交割≈15日, 远月交割≈15日] 与年度贴水窗口的
    重叠日历天数 × 日贴水。windows=[{"per_day":4,"from":"08-01","to":"11-21"}]，窗口每年重复。
    例: 郑棉旧棉 N+1年8/1 起每日贴水4元/吨至11月第15交易日(≈11-21)。"""
    if not windows or not near_month or not far_month:
        return 0.0
    from datetime import date
    try:
        start = date(near_month // 100, near_month % 100, delivery_day)
        end = date(far_month // 100, far_month % 100, delivery_day)
    except Exception:
        return 0.0
    total = 0.0
    for w in windows:
        if "step" in w:            # 阶梯型: 持有期跨过 step 日期即一次性 +amount (每年重复)
            try:
                amt = float(w.get("amount", 0))
                mm, dd = (int(x) for x in str(w["step"]).split("-"))
            except Exception:
                continue
            for y in range(start.year, end.year + 1):
                sd = date(y, mm, dd)
                if start < sd <= end:
                    total += amt
            continue
        try:                       # 每日型: 重叠天数 × per_day
            per = float(w.get("per_day", 0))
            fm, fd = (int(x) for x in str(w["from"]).split("-"))
            tm, td = (int(x) for x in str(w["to"]).split("-"))
        except Exception:
            continue
        for y in range(start.year, end.year + 1):
            ws_ = date(y, fm, fd)
            we_ = date(y + 1, tm, td) if (tm, td) < (fm, fd) else date(y, tm, td)
            o = (min(end, we_) - max(start, ws_)).days
            if o > 0:
                total += per * o
    return round(total, 2)


# --------------------------------------------------------------------------- #
# 合约发现
# --------------------------------------------------------------------------- #
def discover_pairs(api, exchanges, max_months, only_adjacent, products_filter):
    products_filter = set(p.strip() for p in products_filter if p.strip())

    product_months = {}
    for ex in exchanges:
        try:
            syms = api.query_quotes(ins_class="FUTURE", exchange_id=ex, expired=False)
        except Exception as e:
            print(f"[警告] query_quotes FUTURE {ex} 失败: {e}"); continue
        for s in syms or []:
            if is_settle_f(s):
                continue
            prod = product_of(s)
            if products_filter and prod not in products_filter:
                continue
            product_months.setdefault((ex, prod), []).append(s)

    combine_map = {}
    combine_rows = []
    for ex in exchanges:
        if ex not in COMBINE_EXCHANGES:
            continue
        try:
            combos = api.query_quotes(ins_class="COMBINE", exchange_id=ex, expired=False)
        except Exception as e:
            print(f"[警告] query_quotes COMBINE {ex} 失败: {e}"); continue
        for c in combos or []:
            parsed = parse_combine(c)
            if not parsed:
                continue
            _ex, _prefix, f1, f2 = parsed
            prod = product_of(f1)
            if products_filter and prod not in products_filter:
                continue
            combine_map[(f1, f2)] = c
            combine_rows.append((ex, prod, f1, f2, c))

    pairs = []
    seen = set()

    def add_pair(ex, prod, a, b):
        near, far = (a, b) if month_key(a) <= month_key(b) else (b, a)
        key = (near, far)
        if key in seen:
            return
        seen.add(key)
        comb, orient = None, 0
        if (near, far) in combine_map:
            comb, orient = combine_map[(near, far)], +1
        elif (far, near) in combine_map:
            comb, orient = combine_map[(far, near)], -1
        pairs.append({"exchange": ex, "product": prod, "near": near, "far": far,
                      "comb": comb, "comb_orient": orient})

    for (ex, prod), syms in product_months.items():
        ordered = sorted(set(syms), key=month_key)[:max_months]
        for i in range(len(ordered) - 1):
            add_pair(ex, prod, ordered[i], ordered[i + 1])

    if not only_adjacent:
        for ex, prod, f1, f2, c in combine_rows:
            add_pair(ex, prod, f1, f2)

    pairs.sort(key=lambda p: (p["exchange"], p["product"], month_key(p["near"])))
    return pairs


def collect_symbols(pairs):
    syms = set()
    for p in pairs:
        syms.add(p["near"]); syms.add(p["far"])
        if p["comb"]:
            syms.add(p["comb"])
    return sorted(syms)


# --------------------------------------------------------------------------- #
# 标签
# --------------------------------------------------------------------------- #
def compute_tags(pairs, quotes, meta):
    by_prod = {}
    for p in pairs:
        key = (p["exchange"], product_of(p["near"]))
        by_prod.setdefault(key, set()).update([p["near"], p["far"]])

    tags = {}
    for (ex, prod), syms in by_prod.items():
        ordered = sorted(syms, key=month_key)
        main, main_oi = None, -1
        for s in ordered:
            q = quotes.get(s)
            oi = clean(getattr(q, "open_interest", None)) if q is not None else None
            if oi is not None and oi > main_oi:
                main_oi, main = oi, s
        cancel = meta.get(prod.upper(), {}).get("cancel", set())
        near_sym = ordered[0] if ordered else None
        for i, s in enumerate(ordered):
            t = []
            if main is not None and s == main:
                t.append("主力")
            if s == near_sym:
                t.append("近月")
            if (month_key(s) % 100) in cancel:
                t.append("注销月")
            if i >= 1 and (month_key(ordered[i - 1]) % 100) in cancel:
                t.append("注销月后第一个月")
            tags[s] = t
    return tags


# --------------------------------------------------------------------------- #
# 价差计算
# --------------------------------------------------------------------------- #
def build_row(p, quotes, meta, tags, rate):
    qn = quotes.get(p["near"]); qf = quotes.get(p["far"])
    if qn is None or qf is None:
        return None

    n_bid, n_bidv = clean(qn.bid_price1), clean(qn.bid_volume1)
    n_ask, n_askv = clean(qn.ask_price1), clean(qn.ask_volume1)
    n_last = clean(qn.last_price)
    f_bid, f_bidv = clean(qf.bid_price1), clean(qf.bid_volume1)
    f_ask, f_askv = clean(qf.ask_price1), clean(qf.ask_volume1)
    f_last = clean(qf.last_price)

    def sub(a, b):
        return None if (a is None or b is None) else round(a - b, 6)

    def vmin(a, b):
        return None if (a is None or b is None) else int(min(a, b))

    calc_buy_p = sub(n_ask, f_bid); calc_buy_v = vmin(n_askv, f_bidv)     # 买价差(吃对手:近ask-远bid)
    calc_sell_p = sub(n_bid, f_ask); calc_sell_v = vmin(n_bidv, f_askv)   # 卖价差(近bid-远ask)
    calc_last = sub(n_last, f_last)                                       # 合成最新(近last-远last)

    comb_bid_p = comb_bid_v = comb_ask_p = comb_ask_v = comb_last = None
    if p["comb"]:
        qc = quotes.get(p["comb"])
        if qc is not None:
            iv = lambda v: None if v is None else int(v)
            c_bid, c_bidv = clean(qc.bid_price1), iv(clean(qc.bid_volume1))
            c_ask, c_askv = clean(qc.ask_price1), iv(clean(qc.ask_volume1))
            c_last = clean(qc.last_price)
            if p["comb_orient"] >= 0:
                comb_ask_p, comb_ask_v = c_ask, c_askv
                comb_bid_p, comb_bid_v = c_bid, c_bidv
                comb_last = c_last
            else:
                comb_ask_p = None if c_bid is None else round(-c_bid, 6); comb_ask_v = c_bidv
                comb_bid_p = None if c_ask is None else round(-c_ask, 6); comb_bid_v = c_askv
                comb_last = None if c_last is None else round(-c_last, 6)

    # 合并盘口：买价=能卖出价差的最高价；卖价=能买入价差的最低价
    bids = [x for x in (calc_sell_p, comb_bid_p) if x is not None]   # 卖出方向(bid)
    asks = [x for x in (calc_buy_p, comb_ask_p) if x is not None]    # 买入方向(ask)
    merge_bid = max(bids) if bids else None
    merge_ask = min(asks) if asks else None

    # 最新价差 = 合成最新 夹在合并买卖价之间
    last_sp = calc_last
    if last_sp is not None and merge_bid is not None and merge_ask is not None:
        lo, hi = min(merge_bid, merge_ask), max(merge_bid, merge_ask)
        last_sp = round(min(max(last_sp, lo), hi), 6)

    # 套利空间(组合 vs 合成)
    arb_sell_comb = sub(comb_bid_p, calc_buy_p)
    arb_buy_comb = sub(calc_sell_p, comb_ask_p)
    cands = [x for x in (arb_sell_comb, arb_buy_comb) if x is not None]
    arb_best = max(cands) if cands else None

    prod = product_of(p["near"]).upper()
    m = meta.get(prod, {})
    roll = m.get("roll", ""); fee = m.get("fee")
    # 转抛按具体月份对判定：品种可转抛 且 [近月,远月)无注销月 且 近月未进交割月
    if str(roll).strip().upper() == "Y":
        nm, fm = month_key(p["near"]), month_key(p["far"])
        if nm is None or fm is None or not can_roll(nm, fm, m.get("cancel")) \
                or in_delivery_month(nm, prod):
            roll = "N"
    md = month_diff(p["near"], p["far"])

    ratio_spread = (None if (last_sp is None or not n_last)
                    else round(last_sp / n_last * 100, 3))
    ratio_storage = (None if (last_sp is None or not fee)
                     else round(last_sp / (-fee * 30), 3))
    # 转抛比例 = 最新价差 / −(资金成本 + 仓储成本)
    #   资金成本 = 1.1 × 近月价 × 资金利率%/100 × 月份差/12
    #   仓储成本 = 仓储费 × 30 × 月份差
    roll_ratio = None
    dec = decay_cost(month_key(p["near"]), month_key(p["far"]), m.get("decay"))
    if last_sp is not None and md and n_last:
        cap = 1.1 * n_last * (rate / 100.0) * (md / 12.0)
        stor = (fee or 0.0) * 30 * md
        denom = -(cap + stor + dec)          # 含特殊时间贴水(如郑棉旧棉日贴水)
        roll_ratio = round(last_sp / denom, 4) if denom != 0 else None

    return {
        "交易所": p["exchange"], "品种": p["product"], "转抛": roll,
        "近月": code_of(p["near"]), "近月标签": ",".join(tags.get(p["near"], [])),
        "远月": code_of(p["far"]), "远月标签": ",".join(tags.get(p["far"], [])),
        "价差合约": code_of(p["comb"]) if p["comb"] else "",
        "合并买价": merge_bid, "合并卖价": merge_ask, "最新价差": last_sp,
        "价差比例%": ratio_spread, "仓储费比例": ratio_storage, "转抛比例": roll_ratio,
        "近买价": n_bid, "近买量": n_bidv, "近卖价": n_ask, "近卖量": n_askv, "近最新": n_last,
        "远买价": f_bid, "远买量": f_bidv, "远卖价": f_ask, "远卖量": f_askv, "远最新": f_last,
        "合成买价差": calc_buy_p, "合成买量": calc_buy_v,
        "合成卖价差": calc_sell_p, "合成卖量": calc_sell_v, "合成最新": calc_last,
        "组合买价": comb_bid_p, "组合买量": comb_bid_v,
        "组合卖价": comb_ask_p, "组合卖量": comb_ask_v, "组合最新": comb_last,
        "卖组合套利": arb_sell_comb, "买组合套利": arb_buy_comb, "最优套利": arb_best,
        # 隐藏字段供 HTML 实时重算转抛比例
        "_spread": last_sp, "_fee": (fee if fee else 0.0), "_md": md, "_near": n_last, "_dec": dec,
    }


COLUMNS = [
    "交易所", "品种", "转抛", "近月", "近月标签", "远月", "远月标签", "价差合约",
    "合并买价", "合并卖价", "最新价差", "价差比例%", "仓储费比例", "转抛比例",
    "近买价", "近买量", "近卖价", "近卖量", "近最新",
    "远买价", "远买量", "远卖价", "远卖量", "远最新",
    "合成买价差", "合成买量", "合成卖价差", "合成卖量", "合成最新",
    "组合买价", "组合买量", "组合卖价", "组合卖量", "组合最新",
    "卖组合套利", "买组合套利", "最优套利",
]
HIDDEN = ["_spread", "_fee", "_md", "_near", "_dec"]
TEXT_COLS = {"交易所", "品种", "转抛", "近月", "近月标签", "远月", "远月标签", "价差合约"}


def build_table(pairs, quotes, meta, rate=6.0):
    tags = compute_tags(pairs, quotes, meta)
    rows = [r for r in (build_row(p, quotes, meta, tags, rate) for p in pairs) if r]
    return pd.DataFrame(rows, columns=COLUMNS + HIDDEN)


# --------------------------------------------------------------------------- #
# 输出 Excel
# --------------------------------------------------------------------------- #
def write_excel(df, path, ts):
    from openpyxl.utils import get_column_letter
    disp = df[COLUMNS]
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        disp.to_excel(writer, index=False, sheet_name="价差监控", startrow=1)
        ws = writer.sheets["价差监控"]
        ws.cell(row=1, column=1, value=f"更新时间 {ts}    共 {len(disp)} 个价差对")
        ws.freeze_panes = "C3"
        last = get_column_letter(len(disp.columns))
        ws.auto_filter.ref = f"A2:{last}{len(disp) + 2}"
        for i, col in enumerate(disp.columns, start=1):
            ws.column_dimensions[get_column_letter(i)].width = max(8, len(str(col)) * 2 + 2)


def fmt(v):
    if v is None or (isinstance(v, float) and math.isnan(v)):
        return ""
    if isinstance(v, float):
        return f"{v:g}"
    return str(v)


def _num(v):
    return "" if (v is None or (isinstance(v, float) and math.isnan(v))) else f"{v:g}"


# --------------------------------------------------------------------------- #
# 输出 HTML
# --------------------------------------------------------------------------- #
def write_html(df, path, ts, refresh_seconds, default_rate=6.0):
    cols = list(COLUMNS)
    col_idx = {c: i for i, c in enumerate(cols)}
    numeric_idx = [i for i, c in enumerate(cols) if c not in TEXT_COLS]

    head_cells = "".join(
        f"<th data-col='{i}' title='点击排序'>{html.escape(c)}</th>" for i, c in enumerate(cols)
    )
    body_rows = []
    for _, r in df.iterrows():
        best = r["最优套利"]
        cls = " class='arb'" if (isinstance(best, (int, float)) and not is_nan(best) and best > 0) else ""
        data = (f" data-spread=\"{_num(r['_spread'])}\""
                f" data-fee=\"{_num(r['_fee'])}\""
                f" data-md=\"{_num(r['_md'])}\""
                f" data-near=\"{_num(r['_near'])}\""
                f" data-dec=\"{_num(r['_dec'])}\"")
        tds = []
        for c in cols:
            v = r[c]
            align = "left" if c in TEXT_COLS else "right"
            tds.append(f"<td style='text-align:{align}'>{html.escape(fmt(v))}</td>")
        body_rows.append(f"<tr{cls}{data}>{''.join(tds)}</tr>")
    body = "\n".join(body_rows)

    exchanges = sorted(set(str(x) for x in df["交易所"]))
    exch_opts = "".join(f"<option value='{html.escape(e)}'>{html.escape(e)}</option>" for e in exchanges)
    cfg_js = json.dumps({"COL": col_idx, "NUM": numeric_idx, "REFRESH": refresh_seconds,
                         "RATE": default_rate}, ensure_ascii=False)

    doc = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="utf-8">
<title>期货价差套利监控</title>
<style>
  body {{ font-family:"Microsoft YaHei",Consolas,monospace; margin:10px; background:#0f1115; color:#e6e6e6; }}
  h2 {{ margin:0 0 6px; }}
  .bar {{ display:flex; flex-wrap:wrap; gap:8px; align-items:center; margin-bottom:8px; font-size:13px; }}
  .bar label {{ color:#9aa4b2; }}
  select,input {{ background:#1c2230; color:#e6e6e6; border:1px solid #2a2f3a; border-radius:4px; padding:3px 6px; font-size:13px; }}
  input[type=number] {{ width:64px; }}
  .meta {{ color:#9aa4b2; font-size:12px; margin-left:auto; }}
  table {{ border-collapse:collapse; width:100%; font-size:12px; }}
  th,td {{ border:1px solid #2a2f3a; padding:3px 6px; white-space:nowrap; }}
  thead th {{ position:sticky; top:0; background:#1c2230; color:#cfe1ff; cursor:pointer; user-select:none; z-index:2; }}
  thead th.sortasc::after {{ content:" \\25B2"; }}
  thead th.sortdesc::after {{ content:" \\25BC"; }}
  tbody tr:nth-child(even) {{ background:#161a22; }}
  tbody tr:hover {{ background:#23314a; }}
  tr.arb td {{ background:#2d3a1f; }}
  .legend {{ margin-top:8px; color:#9aa4b2; font-size:12px; line-height:1.6; }}
</style>
</head>
<body>
  <h2>期货价差套利监控</h2>
  <div class="bar">
    <label>交易所 <select id="fExch"><option value="">全部</option>{exch_opts}</select></label>
    <label>转抛 <select id="fRoll"><option value="">全部</option><option value="Y">Y</option><option value="N">N</option></select></label>
    <label>标签 <select id="fTag"><option value="">全部</option><option>主力</option><option>注销月</option><option>注销月后第一个月</option><option>近月</option></select></label>
    <label>品种/合约 <input id="fProd" placeholder="如 m / CF / rb2510" size="12"></label>
    <label><input type="checkbox" id="fComb"> 仅有价差合约</label>
    <label>资金利率% <input type="number" id="rate" step="0.1" value="{default_rate}"></label>
    <label><input type="checkbox" id="fAuto"> 自动刷新</label>
    <button id="btnReset">重置</button>
    <span class="meta">更新 {html.escape(ts)} · 共 {len(df)} 对 · <span id="shown"></span></span>
  </div>
  <table id="t">
    <thead><tr>{head_cells}</tr></thead>
    <tbody>
{body}
    </tbody>
  </table>
  <div class="legend">
    方向 <b>近月−远月</b>。合并买价=max(合成卖价差,组合买价)；合并卖价=min(合成买价差,组合卖价)；最新价差=合成最新夹在二者之间。<br>
    价差比例% = 最新价差/近月最新×100；仓储费比例 = 最新价差/(−仓储费×30)；
    转抛比例 = 最新价差 / −(1.1×近月价×资金利率%/100×月份差/12 + 仓储费×30×月份差)。
    改"资金利率%"会实时重算转抛比例。点列头排序、可多条件筛选；状态本地保存。绿行=组合与合成存在正套利。
    转抛列按月份对判定：品种可转抛 且 [近月,远月)区间无注销月 且 近月未进交割月(股指例外)。
  </div>
<script>
const CFG = {cfg_js};
const KEY = "spreadMonitorState";
const def = {{sortCol:null, sortDir:1, fExch:"", fRoll:"", fTag:"", fProd:"", fComb:false, auto:true, rate:CFG.RATE}};
let st = Object.assign({{}}, def);
try {{ Object.assign(st, JSON.parse(localStorage.getItem(KEY)||"{{}}")); }} catch(e) {{}}
let timer = null;

const $ = id => document.getElementById(id);
const rows = () => [...document.querySelectorAll("#t tbody tr")];
const txt = (tr,i) => tr.children[i] ? tr.children[i].textContent.trim() : "";

function recalcRoll() {{
  const rate = parseFloat(st.rate) / 100; const ci = CFG.COL["转抛比例"];
  rows().forEach(tr => {{
    const sp = parseFloat(tr.dataset.spread), fee = parseFloat(tr.dataset.fee)||0,
          md = parseFloat(tr.dataset.md), near = parseFloat(tr.dataset.near),
          dec = parseFloat(tr.dataset.dec)||0;
    let val = "";
    if (!isNaN(sp) && md && !isNaN(near)) {{ const d = -(1.1*near*rate*(md/12) + fee*30*md + dec); if (d !== 0) val = (sp/d).toFixed(4); }}
    tr.children[ci].textContent = val;
  }});
}}
function applyFilter() {{
  const C = CFG.COL, p = st.fProd.toLowerCase(); let shown = 0;
  rows().forEach(tr => {{
    let ok = true;
    if (st.fExch && txt(tr,C["交易所"]) !== st.fExch) ok = false;
    if (st.fRoll && txt(tr,C["转抛"]) !== st.fRoll) ok = false;
    if (st.fTag) {{ const tg = (txt(tr,C["近月标签"])+","+txt(tr,C["远月标签"])).split(","); if (!tg.includes(st.fTag)) ok = false; }}
    if (p) {{ const hay = (txt(tr,C["品种"])+txt(tr,C["近月"])+txt(tr,C["远月"])).toLowerCase(); if (!hay.includes(p)) ok = false; }}
    if (st.fComb && !txt(tr,C["价差合约"])) ok = false;
    tr.style.display = ok ? "" : "none"; if (ok) shown++;
  }});
  $("shown").textContent = "筛选后 " + shown + " 行";
}}
function applySort() {{
  document.querySelectorAll("thead th").forEach(th => th.classList.remove("sortasc","sortdesc"));
  if (st.sortCol == null) return;
  const i = st.sortCol, dir = st.sortDir, num = CFG.NUM.includes(i);
  const th = document.querySelector(`thead th[data-col='${{i}}']`);
  if (th) th.classList.add(dir > 0 ? "sortasc" : "sortdesc");
  const tb = document.querySelector("#t tbody");
  rows().sort((a,b) => {{
    let x = txt(a,i), y = txt(b,i);
    if (num) {{ let nx = parseFloat(x), ny = parseFloat(y); nx = isNaN(nx)?-Infinity:nx; ny = isNaN(ny)?-Infinity:ny; return (nx-ny)*dir; }}
    return x.localeCompare(y, "zh") * dir;
  }}).forEach(r => tb.appendChild(r));
}}
function save() {{ localStorage.setItem(KEY, JSON.stringify(st)); }}
function scheduleRefresh() {{
  if (timer) {{ clearTimeout(timer); timer = null; }}
  if (st.auto) timer = setTimeout(() => location.reload(), CFG.REFRESH * 1000);
}}

document.querySelectorAll("thead th").forEach(th => th.addEventListener("click", () => {{
  const i = +th.dataset.col;
  if (st.sortCol === i) st.sortDir = -st.sortDir; else {{ st.sortCol = i; st.sortDir = 1; }}
  save(); applySort();
}}));
$("fExch").onchange = e => {{ st.fExch = e.target.value; save(); applyFilter(); }};
$("fRoll").onchange = e => {{ st.fRoll = e.target.value; save(); applyFilter(); }};
$("fTag").onchange  = e => {{ st.fTag  = e.target.value; save(); applyFilter(); }};
$("fProd").oninput  = e => {{ st.fProd = e.target.value; save(); applyFilter(); }};
$("fComb").onchange = e => {{ st.fComb = e.target.checked; save(); applyFilter(); }};
$("rate").oninput   = e => {{ st.rate = e.target.value; save(); recalcRoll(); applySort(); }};
$("fAuto").onchange = e => {{ st.auto = e.target.checked; save(); scheduleRefresh(); }};
$("btnReset").onclick = () => {{ st = Object.assign({{}}, def); save(); location.reload(); }};

$("fExch").value = st.fExch; $("fRoll").value = st.fRoll; $("fTag").value = st.fTag;
$("fProd").value = st.fProd; $("fComb").checked = st.fComb; $("fAuto").checked = st.auto;
$("rate").value = st.rate;
recalcRoll(); applyFilter(); applySort(); scheduleRefresh();
</script>
</body>
</html>"""
    with open(path, "w", encoding="utf-8") as fp:
        fp.write(doc)


# --------------------------------------------------------------------------- #
# 主流程
# --------------------------------------------------------------------------- #
def load_config(path):
    cfg = configparser.ConfigParser()
    cfg.read(path, encoding="utf-8")
    g = lambda k, d="": cfg.get("settings", k, fallback=d)
    return dict(
        user=cfg.get("auth", "user", fallback="").strip(),
        password=cfg.get("auth", "password", fallback="").strip(),
        exchanges=[x.strip() for x in g("exchanges", ",".join(FUTURE_EXCHANGES)).split(",") if x.strip()],
        max_months=cfg.getint("settings", "max_months", fallback=6),
        only_adjacent=cfg.getboolean("settings", "only_adjacent", fallback=True),
        products=[x.strip() for x in g("products", "").split(",")],
        refresh_seconds=cfg.getint("settings", "refresh_seconds", fallback=3),
        html_refresh_seconds=cfg.getint("settings", "html_refresh_seconds", fallback=10),
        funding_rate=cfg.getfloat("settings", "funding_rate", fallback=6.0),
        output_dir=g("output_dir", "output").strip(),
        product_config=g("product_config", "all_product_config20260629.xlsx").strip(),
    )


def main():
    here = os.path.dirname(os.path.abspath(__file__))
    cfg = load_config(os.path.join(here, "config.ini"))
    if not cfg["user"] or not cfg["password"]:
        raise SystemExit("请先在 config.ini 的 [auth] 填写快期账户 user / password")

    out_dir = os.path.join(here, cfg["output_dir"]); os.makedirs(out_dir, exist_ok=True)
    html_path = os.path.join(out_dir, "spread_monitor.html")
    xlsx_path = os.path.join(out_dir, "spread_monitor.xlsx")

    meta = load_product_meta(os.path.join(here, cfg["product_config"]))
    print(f"载入品种配置 {len(meta)} 个。")

    print("连接行情服务器 ...")
    api = TqApi(auth=TqAuth(cfg["user"], cfg["password"]))
    try:
        print("发现合约并构造相邻月份配对 ...")
        pairs = discover_pairs(api, cfg["exchanges"], cfg["max_months"],
                               cfg["only_adjacent"], cfg["products"])
        with_comb = sum(1 for p in pairs if p["comb"])
        print(f"共 {len(pairs)} 个价差对，其中 {with_comb} 个有交易所价差合约。")
        if not pairs:
            raise SystemExit("没有发现可监控的合约，请检查 exchanges / products。")

        symbols = collect_symbols(pairs)
        print(f"订阅 {len(symbols)} 个合约盘口 ...")
        qlist = api.get_quote_list(symbols)
        quotes = {q.instrument_id: q for q in qlist}

        last_write = 0.0; n = 0
        while True:
            api.wait_update()
            now = time.time()
            if now - last_write >= cfg["refresh_seconds"]:
                df = build_table(pairs, quotes, meta, cfg["funding_rate"])
                ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                write_html(df, html_path, ts, cfg["html_refresh_seconds"], cfg["funding_rate"])
                write_excel(df, xlsx_path, ts)
                last_write = now; n += 1
                arb = df["最优套利"].apply(lambda x: isinstance(x, (int, float)) and not is_nan(x) and x > 0).sum()
                print(f"[{ts}] 第{n}次刷新  行数={len(df)}  正套利={arb}  -> {html_path}")
    except KeyboardInterrupt:
        print("\n已停止。")
    finally:
        api.close()


if __name__ == "__main__":
    main()
